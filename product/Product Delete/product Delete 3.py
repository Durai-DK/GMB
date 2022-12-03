from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from GMB.Google.Google_login import Google
import time

wb = load_workbook(r"D:\Durai\GMB\product\Data\GMB Product URL.xlsx")
ws = wb.active

driver = webdriver.Chrome(r"D:\Durai\Driver\chromedriver.exe")
g = Google(driver=driver)
g.login()
# l = ["Win a Car*, Bike* & More with Poorvika's Feedbac…",
#      "Free ₹3,999* worth Goodies with Laptops | D…",
#      "Groom & Style with Personal Care Gadgets | …",
#      "Upto 70%* Off on Smart Home Devices this Deep…",
#      "Smart TVs from just ₹9,999*! Grab Free Mi S…",
#      "Flaunt a Brand New Smartwatch | Upto 75%* …",
#      "Snatch Upto ₹7,000* Benefits on Branded Sm…",
#      "Get upto ₹6,000* worth Benefits on Tablets & iP…",
#      "Grab Upto 65% Off on TWS Earbuds! Poorvika …",
#      "Upto 70%* Off on Smart Home Devices this Deep…"]

for r in range(350,460):
    print("")
    print(r)
    time.sleep(5)
    print(ws.cell(row=r,column=2).value)

    try:
        driver.get(url=ws.cell(row=r,column=2).value)
        driver.implicitly_wait(5)
        for r1 in driver.find_elements(By.CLASS_NAME,'zTXfQc'):
            for r2 in r1.find_elements(By.CLASS_NAME,'VfPpkd-ksKsZd-XxIAqe'):
                for r3 in r2.find_elements(By.CLASS_NAME,"bgmvLc"):

                    # for _ in range(0,10):
                    #     print("web link :",r3.text)
                    #     print("list :",l[_])
                    #     if r3.text == l[_]:
                    if r3.text == "Mi TV 5A Pro HD Ready ( 32 inch )":
                        r2.click()
                        time.sleep(2)

                        for dk1 in driver.find_elements(By.CLASS_NAME, "l72iR"):
                            for dk2 in dk1.find_elements(By.CLASS_NAME, "VfPpkd-Bz112c-LgbsSe"):

                                if dk2.text == "delete":
                                    dk2.click()

                                    for lk1 in driver.find_elements(By.CLASS_NAME, "XfpsVe"):
                                        for lk2 in lk1.find_elements(By.CLASS_NAME, "RveJvd"):
                                            # print(lk2.text)

                                            if lk2.text == "Delete":
                                                lk2.click()
                                                print("done")
                                                break
    except:
        pass
driver.quit()
driver.close()
